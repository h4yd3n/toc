"""§13 the spreadsheet upload (2026-09-05). Every section takes what its people actually keep: an Excel workbook or a CSV,
formatted however it was formatted. The flow is preview → mapping → commit: the app reads the workbook, finds the header
row under whatever title rows sit above it, proposes what each column means (the S2 drafter's model when a key is set,
header matching otherwise), shows a sample with what it could not place, and lands nothing until a person says yes.

Sections: S1 roster (a unit path like "B/1-101 ARB" builds the task organization), S3 schedule (events, operations, travel),
S4 supply lines and shipments, S6 systems. The author's decisions: Excel is required (units keep Excel, inconsistently
formatted), the unit path is one column the importer splits, and AI proposes but never commits."""
from __future__ import annotations

import csv
import io
import json
import re
import time
import uuid
from datetime import datetime
from typing import Any, Dict, List, Optional, Tuple

from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession

from shared import settings
from .db_models import EventRow, LocationRow, PersonRow, TeamRow, TripRow
from .imports import _dt, _people_index
from .sections import PACE, SUPPLY_CATEGORIES, SYSTEM_CATEGORIES, ShipmentRow, SupplyRow, SystemRow

# ---------------------------------------------------------------- what each section's sheet can carry

TARGETS: Dict[str, Dict[str, Dict[str, Any]]] = {
    "S1": {
        "name":  {"label": "Name", "required": True, "syn": ["name", "full name", "soldier", "member", "employee", "person"]},
        "last_name": {"label": "Last name", "syn": ["last name", "last", "surname", "family name", "lname"]},
        "first_name": {"label": "First name", "syn": ["first name", "first", "given name", "fname"]},
        "rank":  {"label": "Rank / grade", "syn": ["rank", "grade", "pay grade", "rk"]},
        "unit":  {"label": "Unit (path)", "required": True, "syn": ["unit", "uic", "company", "co", "unit path", "organization", "org", "team", "department", "dept", "section", "assigned unit"]},
        "role":  {"label": "Duty position / MOS", "syn": ["duty", "duty position", "position", "mos", "title", "role", "job", "job title", "aoc"]},
        "phone": {"label": "Phone", "syn": ["phone", "cell", "mobile", "telephone", "phone number", "cell phone"]},
        "email": {"label": "Email", "syn": ["email", "e-mail", "mail", "email address"]},
        "is_vip": {"label": "VIP / key leader", "syn": ["vip", "key leader", "kl", "principal", "executive"]},
        "id":    {"label": "ID", "syn": ["id", "dodid", "edipi", "employee id", "emp id", "person id", "badge"]},
    },
    "S3": {
        "kind":  {"label": "Kind (event / trip / operation)", "syn": ["kind", "type", "category", "activity type"]},
        "name":  {"label": "Name / title", "required": True, "syn": ["name", "title", "event", "activity", "operation", "exercise", "subject", "mission"]},
        "start": {"label": "Start", "required": True, "syn": ["start", "start date", "begin", "from", "date", "depart", "departure", "start time", "dtg start"]},
        "end":   {"label": "End", "syn": ["end", "end date", "finish", "to", "return", "until", "end time", "dtg end", "complete"]},
        "place": {"label": "Place / venue / destination", "syn": ["place", "location", "venue", "destination", "site", "where", "dest", "location name", "ao"]},
        "lat":   {"label": "Latitude", "syn": ["lat", "latitude", "y"]},
        "lon":   {"label": "Longitude", "syn": ["lon", "lng", "long", "longitude", "x"]},
        "who":   {"label": "Who (traveler / lead)", "syn": ["who", "traveler", "person", "name", "lead", "oic", "attendee", "poc", "email"]},
        "purpose": {"label": "Purpose / description", "syn": ["purpose", "description", "notes", "remarks", "desc", "details", "task"]},
    },
    "S4": {
        "site":  {"label": "Site / unit", "syn": ["site", "location", "unit", "where", "facility", "post", "base", "fob", "company", "battalion", "org"]},
        "category": {"label": "Class / category", "syn": ["class", "category", "class of supply", "cos", "type", "commodity"]},
        "item":  {"label": "Item", "required": True, "syn": ["item", "nomenclature", "description", "name", "supply", "equipment", "lin", "nsn", "material"]},
        "on_hand": {"label": "On hand", "required": True, "syn": ["on hand", "oh", "qty", "quantity", "current", "have", "available", "count", "stock", "onhand", "o/h", "fmc"]},
        "required": {"label": "Required / authorized", "syn": ["required", "req", "authorized", "auth", "needed", "target", "minimum", "min", "reorder", "should have", "assigned"]},
        "unit_of_measure": {"label": "Unit of measure", "syn": ["uom", "unit of measure", "unit of issue", "ui", "measure", "units"]},
        "note":  {"label": "Note", "syn": ["note", "notes", "remarks", "comment", "comments", "status"]},
        # shipments share the sheet kind selector
        "description": {"label": "Shipment description", "syn": ["shipment", "cargo", "load", "consignment"]},
        "quantity": {"label": "Quantity (text)", "syn": ["quantity", "qty text", "amount"]},
        "from": {"label": "From", "syn": ["from", "origin", "shipper", "source"]},
        "to":   {"label": "To (site)", "syn": ["to", "destination", "consignee", "deliver to"]},
        "eta":  {"label": "ETA", "syn": ["eta", "arrival", "due", "expected", "delivery date", "rdd"]},
        "status": {"label": "Status", "syn": ["status", "state"]},
        "priority": {"label": "Priority", "syn": ["priority", "pri", "urgency"]},
        "ref":  {"label": "Reference", "syn": ["ref", "reference", "tcn", "tracking", "document", "doc", "conf", "confirmation"]},
    },
    "S6": {
        "site":  {"label": "Site / command post", "syn": ["site", "location", "cp", "command post", "toc", "node", "where", "unit"]},
        "name":  {"label": "System / net", "required": True, "syn": ["system", "name", "net", "network", "service", "circuit", "asset", "equipment"]},
        "category": {"label": "Category", "syn": ["category", "type", "kind", "class"]},
        "pace":  {"label": "PACE role", "syn": ["pace", "pace role", "role", "primary/alternate", "p/a/c/e"]},
        "status": {"label": "Status (up / degraded / down)", "required": True, "syn": ["status", "state", "condition", "up/down", "opstat", "operational"]},
        "note":  {"label": "Note", "syn": ["note", "notes", "remarks", "comment", "comments", "issue"]},
    },
}
S4_KINDS = {"supply": ("site", "category", "item", "on_hand", "required", "unit_of_measure", "note"),
            "shipments": ("description", "category", "quantity", "from", "to", "eta", "status", "priority", "ref", "note")}


# ---------------------------------------------------------------- reading the workbook, whatever shape it is in

def _cell(v: Any) -> str:
    if v is None: return ""
    if isinstance(v, datetime): return v.strftime("%Y-%m-%dT%H:%M:%S")
    if isinstance(v, float) and v.is_integer(): return str(int(v))
    return str(v).strip()


def read_workbook(data: bytes, filename: str) -> Dict[str, List[List[str]]]:
    """Every sheet as a grid of strings. .xlsx / .xlsm via openpyxl (values, not formulas); .csv as one sheet."""
    name = (filename or "").lower()
    if name.endswith((".xlsx", ".xlsm", ".xltx")):
        from openpyxl import load_workbook
        wb = load_workbook(io.BytesIO(data), data_only=True, read_only=True)
        out = {}
        for ws in wb.worksheets:
            rows = [[_cell(c) for c in r] for r in ws.iter_rows(values_only=True)]
            rows = [r for r in rows if any(x for x in r)]
            if rows: out[ws.title] = rows
        return out
    text = data.decode("utf-8-sig", errors="replace")
    rows = [[(c or "").strip() for c in r] for r in csv.reader(io.StringIO(text))]
    return {"Sheet1": [r for r in rows if any(r)]}


def find_header(rows: List[List[str]]) -> int:
    """The header is the first row that is mostly short text and has something under it — title rows and blank rows sit above it."""
    best, best_score = 0, -1.0
    for i, r in enumerate(rows[:15]):
        cells = [c for c in r if c]
        if len(cells) < 2: continue
        texty = sum(1 for c in cells if not re.fullmatch(r"[-+]?\d[\d,.]*", c) and len(c) <= 40)
        width = len(cells)
        below = rows[i + 1] if i + 1 < len(rows) else []
        score = texty / max(width, 1) + min(width, 12) / 12 + (0.5 if any(below) else 0)
        if score > best_score: best, best_score = i, score
    return best


def sheet_table(rows: List[List[str]], header_row: Optional[int] = None) -> Tuple[List[str], List[Dict[str, str]]]:
    h = find_header(rows) if header_row is None else header_row
    header = [c or f"col{i + 1}" for i, c in enumerate(rows[h])]
    seen: Dict[str, int] = {}
    cols = []
    for c in header:
        k = c
        while k in seen: seen[k] = seen[k] + 1; k = f"{c} ({seen[c]})"
        seen.setdefault(k, 1); cols.append(k)
    data = [{cols[i]: (r[i] if i < len(r) else "") for i in range(len(cols))} for r in rows[h + 1:] if any(r)]
    return cols, data


# ---------------------------------------------------------------- proposing what each column means

def _norm(s: str) -> str:
    return re.sub(r"[^a-z0-9/ ]+", " ", s.lower()).strip()


def heuristic_mapping(section: str, columns: List[str], samples: List[Dict[str, str]]) -> Dict[str, Optional[str]]:
    targets = TARGETS[section]
    mapping: Dict[str, Optional[str]] = {}
    used = set()
    for col in columns:
        n = _norm(col)
        best, best_len = None, 0
        for t, spec in targets.items():
            for syn in [t.replace("_", " ")] + spec["syn"]:
                if (n == syn or n.startswith(syn + " ") or n.endswith(" " + syn) or (len(syn) > 3 and syn in n)) and len(syn) > best_len and t not in used:
                    best, best_len = t, len(syn)
        # a column of emails or phones is one whatever it is called
        vals = [r.get(col, "") for r in samples if r.get(col)]
        if not best and vals:
            if all("@" in v for v in vals[:5]) and "email" in targets and "email" not in used: best = "email"
            elif all(re.fullmatch(r"[+\d][\d\s().-]{6,}", v) for v in vals[:5]) and "phone" in targets and "phone" not in used: best = "phone"
        mapping[col] = best
        if best: used.add(best)
    return mapping


async def model_mapping(section: str, columns: List[str], samples: List[Dict[str, str]]) -> Optional[Dict[str, Optional[str]]]:
    """The S2 drafter's model reads the headers and a few rows and says what each column is. None without a key or on any trouble."""
    if not settings.get("ANTHROPIC_API_KEY"):
        return None
    try:
        from anthropic import AsyncAnthropic
        client = AsyncAnthropic(api_key=settings.get("ANTHROPIC_API_KEY"))
        targets = {t: spec["label"] for t, spec in TARGETS[section].items()}
        prompt = ("A staff section uploaded a spreadsheet. Map each column name to one of the targets, or null if it fits none. "
                  "Answer with JSON only: {\"mapping\": {\"<column>\": \"<target or null>\"}}.\n\n"
                  f"Section: {section}\nTargets: {json.dumps(targets)}\nColumns: {json.dumps(columns)}\nSample rows: {json.dumps(samples[:5])}")
        resp = await client.messages.create(model=settings.get("TOC_MODEL") or "claude-opus-5", max_tokens=1024, messages=[{"role": "user", "content": prompt}])
        text = "".join(getattr(b, "text", "") for b in resp.content if getattr(b, "type", "") == "text")
        data = json.loads(text[text.find("{"): text.rfind("}") + 1]).get("mapping", {})
        return {c: (data.get(c) if data.get(c) in TARGETS[section] else None) for c in columns}
    except Exception:
        return None


# ---------------------------------------------------------------- the pending uploads (preview → commit)

_pending: Dict[str, Dict[str, Any]] = {}


def remember(payload: Dict[str, Any]) -> str:
    now = time.time()
    for k in [k for k, v in _pending.items() if now - v["at"] > 1800]:
        _pending.pop(k, None)
    uid = f"up_{uuid.uuid4().hex[:10]}"
    _pending[uid] = {"at": now, **payload}
    return uid


def recall(upload_id: str) -> Optional[Dict[str, Any]]:
    return _pending.get(upload_id)


async def preview(section: str, data: bytes, filename: str, sheet: Optional[str] = None, header_row: Optional[int] = None) -> Dict[str, Any]:
    book = read_workbook(data, filename)
    if not book:
        raise ValueError("no rows found in the file")
    sheet = sheet if sheet in book else next(iter(book))
    cols, rows = sheet_table(book[sheet], header_row)
    samples = rows[:8]
    mapping = await model_mapping(section, cols, samples)
    proposed_by = "model" if mapping else "headers"
    if not mapping:
        mapping = heuristic_mapping(section, cols, samples)
    issues = []
    for t, spec in TARGETS[section].items():
        if spec.get("required") and t not in mapping.values() and not (section == "S1" and t == "name" and "last_name" in mapping.values()):
            issues.append(f"no column for {spec['label']}")
    upload_id = remember({"section": section, "filename": filename, "book": book})
    kind = "shipments" if section == "S4" and any(mapping.get(c) in ("description", "eta", "from") for c in cols) else "supply"
    return {"upload_id": upload_id, "section": section, "filename": filename, "sheets": list(book), "sheet": sheet, "header_row": find_header(book[sheet]) if header_row is None else header_row,
            "columns": cols, "rows": len(rows), "samples": samples, "mapping": mapping, "proposed_by": proposed_by, "kind": kind,
            "targets": {t: spec["label"] for t, spec in TARGETS[section].items()}, "issues": issues}


# ---------------------------------------------------------------- landing the rows

def _split_unit(path: str) -> List[Tuple[str, str]]:
    """'B/1-101 ARB' → [(battalion '1-101 ARB'), (company 'B')]; 'HHC/CAB' → [(brigade 'CAB'), (company 'HHC')]; 'A Co, 2-17 CAV' handled the same.
    The rightmost part is the parent. Returns [(echelon, short), ...] from the top down."""
    parts = [p.strip() for p in re.split(r"[/,]", path) if p.strip()]
    if not parts: return []
    if len(parts) == 1:
        return [("company", parts[0])]
    parent, child = parts[-1], parts[0]
    ech = "brigade" if re.search(r"\b(bde|brigade|cab|bct|group|grp|regiment|regt|division|div|corps)\b", parent, re.I) else "battalion"
    return [(ech, parent), ("company", child)]


async def _team_for_unit(session: AsyncSession, path: str, teams: Dict[str, TeamRow], root_loc: str) -> TeamRow:
    """Find or build the team for a unit path, hanging new ones under the existing root or a new brigade-level root."""
    chain = _split_unit(path)
    by_short = {(t.short or "").lower(): t for t in teams.values() if t.short}
    by_name = {t.name.lower(): t for t in teams.values()}
    root = next((t for t in teams.values() if not t.parent_id and t.echelon in ("brigade", "organization")), None)
    parent = root
    for i, (ech, short) in enumerate(chain):
        key = short.lower()
        t = None
        if parent is not None and ech == "company":  # a company is found under its parent by its letter: "A" matches "A/1", "A/1 ATK", "A Co"
            t = next((c for c in teams.values() if c.parent_id == parent.id and ((c.short or "").lower().split("/")[0].strip() == key or c.name.lower().startswith(key + " ") or c.name.lower() == key)), None)
        if t is None:
            t = by_short.get(key) or by_name.get(key) or next((c for c in teams.values() if (c.short or "").lower().replace(" ", "") == key.replace(" ", "")), None)
        if t is None:
            if ech in ("brigade", "battalion") and parent is None and root is None:
                t = TeamRow(id=f"t_{uuid.uuid4().hex[:6]}", name=short, location_id=root_loc, function="hq", is_security=False, parent_id=None, echelon=ech, short=short, equipment=None)
                root = t
            else:
                p = parent or root
                if p is None:  # a bare company with no root: make a root to hang it on
                    p = TeamRow(id=f"t_{uuid.uuid4().hex[:6]}", name="Task Organization", location_id=root_loc, function="hq", is_security=False, parent_id=None, echelon="brigade", short="TF", equipment=None)
                    session.add(p); teams[p.id] = p; by_short["tf"] = p; root = p
                t = TeamRow(id=f"t_{uuid.uuid4().hex[:6]}", name=(f"{short}, {p.short or p.name}" if ech == "company" else short), location_id=p.location_id, function="hq" if short.lower().startswith("hhc") else "line",
                            is_security=False, parent_id=p.id, echelon=ech, short=(f"{short}/{p.short}" if ech == "company" and p.echelon != "brigade" and p.short else short), equipment=None)
            session.add(t); teams[t.id] = t; by_short[(t.short or "").lower()] = t; by_name[t.name.lower()] = t
        parent = t
    return parent  # the leaf


def _bool(v: str) -> bool:
    return (v or "").strip().lower() in ("1", "true", "yes", "y", "x", "vip", "★", "*")


def _num(v: str) -> Optional[float]:
    try: return float(str(v).replace(",", "").strip())
    except (TypeError, ValueError): return None


def _locate(name: str, locs: List[LocationRow], teams: Dict[str, TeamRow]) -> Optional[LocationRow]:
    n = (name or "").strip().lower()
    if not n: return None
    for l in locs:
        if l.name.lower() == n or l.id.lower() == n: return l
    for l in locs:
        if n in l.name.lower() or l.name.lower() in n: return l
    for t in teams.values():  # a unit name places its supplies at the unit's site
        if t.name.lower() == n or (t.short or "").lower() == n:
            return next((l for l in locs if l.id == t.location_id), None)
    return None


async def commit(session: AsyncSession, upload_id: str, sheet: str, mapping: Dict[str, Optional[str]], actor: str, kind: str = "supply", source: Optional[str] = None) -> Dict[str, Any]:
    pend = recall(upload_id)
    if not pend:
        raise KeyError("upload expired — preview again")
    section, book = pend["section"], pend["book"]
    if sheet not in book:
        raise ValueError("unknown sheet")
    cols, rows = sheet_table(book[sheet])
    src = source or f"upload:{pend['filename']}"
    col_of = {t: c for c, t in mapping.items() if t}
    def g(r: Dict[str, str], t: str) -> str: return (r.get(col_of[t], "") if t in col_of else "").strip()
    created = updated = skipped = 0; errors: List[str] = []
    locs = list((await session.execute(select(LocationRow))).scalars())
    teams = {t.id: t for t in (await session.execute(select(TeamRow))).scalars()}
    root_loc = next((t.location_id for t in teams.values() if not t.parent_id), locs[0].id if locs else None)
    if section == "S1":
        by_id, by_email = await _people_index(session)
        by_name = {p.name.lower(): p for p in by_id.values()}
        for i, r in enumerate(rows, 1):
            name = g(r, "name") or " ".join(x for x in (g(r, "first_name"), g(r, "last_name")) if x)
            unit = g(r, "unit")
            if not name or not unit:
                skipped += 1; errors.append(f"row {i}: needs a name and a unit"); continue
            rank = g(r, "rank")
            if rank and not name.upper().startswith(rank.upper()): name = f"{rank} {name}"
            team = await _team_for_unit(session, unit, teams, root_loc)
            p = by_id.get(g(r, "id")) or by_email.get(g(r, "email").lower()) or by_name.get(name.lower())
            if p:
                p.name, p.team_id, p.role, p.source = name, team.id, g(r, "role") or p.role, src
                if g(r, "phone"): p.phone = g(r, "phone")
                if g(r, "email"): p.email = g(r, "email")
                if "is_vip" in col_of: p.is_vip = _bool(g(r, "is_vip"))
                updated += 1
            else:
                p = PersonRow(id=g(r, "id") or f"p_up_{uuid.uuid4().hex[:6]}", name=name, role=g(r, "role"), team_id=team.id, is_vip=_bool(g(r, "is_vip")), phone=g(r, "phone") or None, email=g(r, "email") or None, source=src)
                session.add(p); by_id[p.id] = p; by_name[name.lower()] = p
                if p.email: by_email[p.email.lower()] = p
                created += 1
    elif section == "S3":
        by_id, by_email = await _people_index(session)
        by_name = {p.name.lower(): p for p in by_id.values()}
        for i, r in enumerate(rows, 1):
            name, st, en = g(r, "name"), _dt(g(r, "start")), _dt(g(r, "end")) or _dt(g(r, "start"))
            k = (g(r, "kind") or "").lower()
            k = "trip" if "trip" in k or "travel" in k or "tdy" in k else ("operation" if "op" in k else "event")
            if not name or not st:
                skipped += 1; errors.append(f"row {i}: needs a name and a start"); continue
            lat, lon = _num(g(r, "lat")), _num(g(r, "lon"))
            loc = _locate(g(r, "place"), locs, teams)
            if lat is None or lon is None:
                if loc: lat, lon = loc.lat, loc.lon
                else:
                    skipped += 1; errors.append(f"row {i}: '{g(r, 'place') or name}' is not a known place and has no lat/lon"); continue
            if k == "trip":
                who = g(r, "who"); p = by_email.get(who.lower()) or by_name.get(who.lower()) or by_id.get(who)
                if not p:
                    skipped += 1; errors.append(f"row {i}: traveler '{who}' is not in the directory"); continue
                home = teams[p.team_id].location_id if p.team_id in teams else root_loc
                session.add(TripRow(id=f"trip_up_{uuid.uuid4().hex[:6]}", person_id=p.id, origin_location_id=home, dest_location_id=loc.id if loc else None, dest_name=g(r, "place") or name,
                                    dest_lat=lat, dest_lon=lon, depart_at=st, return_at=en if en and en > st else st, purpose=g(r, "purpose") or name, created_by=actor, source=src))
            else:
                session.add(EventRow(id=f"evt_up_{uuid.uuid4().hex[:6]}", name=name, event_type="operation" if k == "operation" else "event", venue_location_id=loc.id if loc else None,
                                     venue_name=g(r, "place") or (loc.name if loc else name), venue_lat=lat, venue_lon=lon, start_at=st, end_at=en if en and en > st else st, description=g(r, "purpose"), source=src))
            created += 1
    elif section == "S4" and kind == "supply":
        existing = list((await session.execute(select(SupplyRow))).scalars())
        for i, r in enumerate(rows, 1):
            item, on, req = g(r, "item"), _num(g(r, "on_hand")), _num(g(r, "required"))
            if not item or on is None:
                skipped += 1; errors.append(f"row {i}: needs an item and an on-hand number"); continue
            loc = _locate(g(r, "site"), locs, teams)
            cat = (g(r, "category") or "other").lower()
            cat = next((c for c in SUPPLY_CATEGORIES if c in cat), {"i": "rations", "iii": "fuel", "v": "ammunition", "viii": "medical", "ix": "parts", "vii": "equipment", "ii": "equipment"}.get(re.sub(r"class\s*", "", cat).strip(), "other"))
            row = next((x for x in existing if x.item.lower() == item.lower() and x.location_id == (loc.id if loc else None)), None)
            if row:
                row.on_hand, row.required, row.unit, row.note, row.updated_by, row.updated_at, row.source = on, req if req is not None else row.required, g(r, "unit_of_measure") or row.unit, g(r, "note") or row.note, actor, datetime.utcnow(), src
                updated += 1
            else:
                session.add(SupplyRow(id=f"sup_up_{uuid.uuid4().hex[:6]}", location_id=loc.id if loc else None, category=cat, item=item, on_hand=on, required=req if req is not None else on,
                                      unit=g(r, "unit_of_measure") or "ea", note=g(r, "note"), updated_by=actor, updated_at=datetime.utcnow(), source=src)); created += 1
    elif section == "S4":
        for i, r in enumerate(rows, 1):
            desc, eta = g(r, "description") or g(r, "item"), _dt(g(r, "eta"))
            if not desc or not eta:
                skipped += 1; errors.append(f"row {i}: needs a description and an ETA"); continue
            loc = _locate(g(r, "to"), locs, teams)
            stt = (g(r, "status") or "planned").lower(); stt = "in_transit" if "transit" in stt or "moving" in stt or "route" in stt else "delayed" if "delay" in stt or "late" in stt else "arrived" if "arriv" in stt or "deliver" in stt else "planned"
            pri = (g(r, "priority") or "routine").lower(); pri = "urgent" if "urg" in pri or "999" in pri or "immediate" in pri else "priority" if "pri" in pri or "high" in pri else "routine"
            session.add(ShipmentRow(id=f"shp_up_{uuid.uuid4().hex[:6]}", description=desc, category=(g(r, "category") or "other").lower() if (g(r, "category") or "other").lower() in SUPPLY_CATEGORIES else "other", quantity=g(r, "quantity"),
                                    from_name=g(r, "from"), to_location_id=loc.id if loc else None, to_name=g(r, "to") or (loc.name if loc else ""), eta=eta, status=stt, priority=pri, carrier="", ref=g(r, "ref") or None,
                                    note=g(r, "note"), updated_by=actor, updated_at=datetime.utcnow(), source=src)); created += 1
    elif section == "S6":
        existing = list((await session.execute(select(SystemRow))).scalars())
        for i, r in enumerate(rows, 1):
            name, stt = g(r, "name"), (g(r, "status") or "").lower()
            if not name or not stt:
                skipped += 1; errors.append(f"row {i}: needs a system and a status"); continue
            status = "down" if any(w in stt for w in ("down", "out", "nmc", "red", "off", "fail")) else "degraded" if any(w in stt for w in ("degrad", "amber", "partial", "pmc", "intermittent")) else "up"
            loc = _locate(g(r, "site"), locs, teams)
            pace = next((p for p in PACE if (g(r, "pace") or "").lower().startswith(p[0])), None) if g(r, "pace") else None
            cat = (g(r, "category") or "").lower(); cat = next((c for c in SYSTEM_CATEGORIES if c in cat), "comms" if pace else "other")
            row = next((x for x in existing if x.name.lower() == name.lower() and x.location_id == (loc.id if loc else None)), None)
            if row:
                if row.status != status: row.since = datetime.utcnow()
                row.status, row.pace, row.note, row.updated_by, row.updated_at, row.source = status, pace or row.pace, g(r, "note") or row.note, actor, datetime.utcnow(), src
                updated += 1
            else:
                session.add(SystemRow(id=f"sys_up_{uuid.uuid4().hex[:6]}", name=name, category=cat, location_id=loc.id if loc else None, pace=pace, status=status, since=datetime.utcnow(),
                                      note=g(r, "note"), updated_by=actor, updated_at=datetime.utcnow(), source=src)); created += 1
    await session.commit()
    return {"section": section, "sheet": sheet, "source": src, "created": created, "updated": updated, "skipped": skipped, "errors": errors[:25]}
